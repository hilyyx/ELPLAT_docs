import io
import openpyxl
from openpyxl import load_workbook
from PIL import Image
import numpy as np
import cv2


def decode_image_bytes(image_data):
    """
    Принимает бинарные данные картинки, переводит на белый фон,
    масштабирует и пытается распознать QR-код через OpenCV.
    """
    try:
        # Открываем изображение через Pillow
        pil_img = Image.open(io.BytesIO(image_data))
        img_np = np.array(pil_img)

        # Переводим на белый фон, если есть альфа-канал (прозрачность)
        if len(img_np.shape) >= 3 and img_np.shape[2] == 4:
            alpha = img_np[:, :, 3] / 255.0
            alpha_3d = np.dstack([alpha, alpha, alpha])
            rgb_part = img_np[:, :, :3]

            white_bg = np.ones_like(rgb_part) * 255
            composited = (rgb_part * alpha_3d + white_bg * (1.0 - alpha_3d)).astype(np.uint8)
        else:
            composited = img_np

        # Переводим в оттенки серого
        gray = cv2.cvtColor(composited, cv2.COLOR_RGB2GRAY)

        # Масштабируем до оптимального для OpenCV размера 400x400
        resized = cv2.resize(gray, (400, 400), interpolation=cv2.INTER_NEAREST)

        # Добавляем 30 пикселей белых полей со всех сторон (Quiet Zone)
        padded = cv2.copyMakeBorder(resized, 30, 30, 30, 30, cv2.BORDER_CONSTANT, value=255)

        # Распознаем через встроенный OpenCV QRCodeDetector
        detector = cv2.QRCodeDetector()
        qr_text, bbox, straight_qrcode = detector.detectAndDecode(padded)

        # Запасной вариант с чуть большим разрешением (на случай мелких деталей)
        if not qr_text:
            resized_backup = cv2.resize(gray, (500, 500), interpolation=cv2.INTER_NEAREST)
            padded_backup = cv2.copyMakeBorder(resized_backup, 45, 45, 45, 45, cv2.BORDER_CONSTANT, value=255)
            qr_text, bbox, straight_qrcode = detector.detectAndDecode(padded_backup)

        if not qr_text:
            return None

        # Исправляем кодировку кириллицы (перевод из латиницы в Windows-1251)
        try:
            qr_text = qr_text.encode('latin1').decode('cp1251')
        except Exception:
            pass

        return qr_text

    except Exception as e:
        print(f"Ошибка при обработке изображения: {e}")
        return None


def parse_gost_qr(qr_text):
    """
    Разбивает строку ГОСТ-платежа на понятный словарь ключ-значение.
    """
    if not qr_text:
        return {}
    parts = qr_text.split('|')
    parsed_data = {'Format': parts[0]}
    for part in parts[1:]:
        if '=' in part:
            key, val = part.split('=', 1)
            parsed_data[key] = val
    return parsed_data


def process_all_qrs(source_excel_path, target_excel_path):
    """
    Основная функция: проходит по всем картинкам в источнике,
    распознает их и обновляет строки в итоговом файле.
    """
    # 1. Загружаем исходный файл с платежками
    try:
        wb_source = load_workbook(source_excel_path)
        sheet_source = wb_source.active
    except FileNotFoundError:
        print(f"Ошибка: Не найден исходный файл '{source_excel_path}'.")
        return

    # Проверяем, есть ли вообще изображения в файле
    total_images = len(sheet_source._images)
    if total_images == 0:
        print("В исходном файле не найдено ни одного изображения.")
        return

    print(f"Найдено изображений в исходном файле: {total_images}. Начинаем распознавание...")

    # 2. Открываем целевой файл для записи данных
    try:
        wb_target = load_workbook(target_excel_path)
        ws_target = wb_target.active
    except FileNotFoundError:
        print(f"Ошибка: Не найден итоговый файл '{target_excel_path}'.")
        return

    # Собираем карту ЛС -> Номер строки из целевого файла для быстрого поиска
    # Это сильно ускорит работу скрипта при больших объемах данных
    ls_to_row_map = {}
    for row in range(2, ws_target.max_row + 1):
        cell_val = ws_target.cell(row=row, column=1).value
        if cell_val is not None:
            clean_ls = str(cell_val).strip()
            ls_to_row_map[clean_ls] = row

    # Статистика обработки
    success_count = 0
    failed_decode_count = 0
    not_found_in_target_count = 0

    # 3. Перебираем все картинки последовательно
    for index, img_obj in enumerate(sheet_source._images, start=1):
        image_data = img_obj._data()

        # Пробуем распознать QR-код
        qr_text = decode_image_bytes(image_data)

        if not qr_text:
            print(
                f"[{index}/{total_images}] Не удалось распознать QR-код (возможно, это логотип или картинка повреждена). Пропуск.")
            failed_decode_count += 1
            continue

        # Разбираем данные ГОСТ-платежа
        details = parse_gost_qr(qr_text)
        target_ls = details.get('persAcc', '').strip()

        if not target_ls:
            print(
                f"[{index}/{total_images}] В QR-коде успешно распознан текст, но отсутствует поле 'persAcc' (Лицевой счет).")
            failed_decode_count += 1
            continue

        row_index = ls_to_row_map.get(target_ls)

        if row_index is None:
            print(
                f"[{index}/{total_images}] Предупреждение: Лицевой счет {target_ls} отсутствует в итоговом файле '{target_excel_path}'.")
            not_found_in_target_count += 1
            continue

        purpose = details.get('Purpose', '')
        group = "гр. 1"
        if "гр. " in purpose:
            try:
                parts = purpose.split("гр. ")
                group = "гр. " + parts[1].strip().split()[0]
            except Exception:
                pass

        child_fio = details.get('childFio', '')
        fi_short = child_fio
        if child_fio:
            fio_parts = child_fio.split()
            if len(fio_parts) >= 2:
                fi_short = f"{fio_parts[0]} {fio_parts[1]}"

        rayon = "Устиновский район"
        kod_uslugi = f"{details.get('PayeeINN', '')}_1"

        cbc = details.get('CBC', '')
        kbk_short = cbc[-4:] if len(cbc) >= 4 else ""

        oktmo = details.get('OKTMO', '')

        sum_raw = details.get('Sum', '0')
        sum_rub = float(sum_raw) / 100 if sum_raw.isdigit() else 0.0

        # --- Запись в ячейки найденной строки ---
        ws_target.cell(row=row_index, column=3, value=group)  # Столбец C: Группа
        ws_target.cell(row=row_index, column=4, value=fi_short)  # Столбец D: ФИ
        ws_target.cell(row=row_index, column=5, value=group)  # Столбец E: Группа (повтор)
        ws_target.cell(row=row_index, column=6, value=rayon)  # Столбец F: Район
        ws_target.cell(row=row_index, column=7, value=kod_uslugi)  # Столбец G: Код услуги
        ws_target.cell(row=row_index, column=8, value=kbk_short)  # Столбец H: Код (4 цифры КБК)
        ws_target.cell(row=row_index, column=9, value=oktmo)  # Столбец I: Окато

        cell_sum = ws_target.cell(row=row_index, column=10, value=sum_rub)
        cell_sum.number_format = '0.00'

        print(f"[{index}/{total_images}] Успешно обработан ЛС: {target_ls} (строка {row_index} обновлена)")
        success_count += 1

    wb_target.save(target_excel_path)

    print("\n" + "=" * 40)
    print("ОБРАБОТКА ЗАВЕРШЕНА!")
    print(f"Всего изображений обработано: {total_images}")
    print(f"Успешно занесено в отчет:     {success_count}")
    print(f"Не удалось распознать (QR):   {failed_decode_count}")
    print(f"ЛС не найден в базе отчета:   {not_found_in_target_count}")
    print("=" * 40)


if __name__ == '__main__':
    source_file = 'Форма № ПД-4 с QR-кодом.xlsx'
    target_file = 'Итог.xlsx'
    process_all_qrs(source_file, target_file)